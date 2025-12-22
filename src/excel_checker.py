"""Excel content checker module."""

import logging
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

logger = logging.getLogger(__name__)


@dataclass
class CheckResult:
    """Check result data class.

    Attributes:
        is_ok: True if all checks passed, False otherwise.
        errors: List of error messages.
    """

    is_ok: bool
    errors: list[str]


def _is_cell_empty(value: str | int | float | None) -> bool:
    """Check if cell value is empty.

    Args:
        value: Cell value.

    Returns:
        bool: True if empty, False otherwise.
    """
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def check_review_record(file_path: Path) -> CheckResult:
    """Check review record Excel file content.

    Args:
        file_path: Path to the review record file.

    Returns:
        CheckResult: Check result containing status and errors.
    """
    errors: list[str] = []

    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # AE2: レビュー名称 (required)
        if _is_cell_empty(ws["AE2"].value):
            errors.append("AE2 (レビュー名称) が未記入です")

        # AE8: 内部メンバ (required)
        if _is_cell_empty(ws["AE8"].value):
            errors.append("AE8 (内部メンバ) が未記入です")

        # AE7: 外部メンバ (required only for 社外 files)
        if "(社外)" in file_path.name:
            if _is_cell_empty(ws["AE7"].value):
                errors.append("AE7 (外部メンバ) が未記入です")

        wb.close()

    except PermissionError:
        logger.error(f"ファイルが開かれているため読み込めません: {file_path}")
        errors.append(f"ファイルアクセスエラー: {file_path.name}")
    except FileNotFoundError:
        logger.error(f"ファイルが見つかりません: {file_path}")
        errors.append(f"ファイルが見つかりません: {file_path.name}")
    except Exception as e:
        logger.error(f"Excelファイルの読み込みに失敗しました: {file_path}, エラー: {e}")
        errors.append(f"Excelファイルの読み込みエラー: {file_path.name}")

    return CheckResult(is_ok=len(errors) == 0, errors=errors)


def check_review_checklist(file_path: Path) -> CheckResult:
    """Check review checklist Excel file content.

    Args:
        file_path: Path to the review checklist file.

    Returns:
        CheckResult: Check result containing status and errors.
    """
    errors: list[str] = []

    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # E4-E6: 案件情報 (required)
        if _is_cell_empty(ws["E4"].value):
            errors.append("E4 (案件情報) が未記入です")
        if _is_cell_empty(ws["E5"].value):
            errors.append("E5 (案件情報) が未記入です")
        if _is_cell_empty(ws["E6"].value):
            errors.append("E6 (案件情報) が未記入です")

        # N6: 日付 (required)
        if _is_cell_empty(ws["N6"].value):
            errors.append("N6 (日付) が未記入です")

        # M15: 外部メンバ (required only for 社外 files)
        if "社外" in file_path.name:
            if _is_cell_empty(ws["M15"].value):
                errors.append("M15 (外部メンバ) が未記入です")

        wb.close()

    except PermissionError:
        logger.error(f"ファイルが開かれているため読み込めません: {file_path}")
        errors.append(f"ファイルアクセスエラー: {file_path.name}")
    except FileNotFoundError:
        logger.error(f"ファイルが見つかりません: {file_path}")
        errors.append(f"ファイルが見つかりません: {file_path.name}")
    except Exception as e:
        logger.error(f"Excelファイルの読み込みに失敗しました: {file_path}, エラー: {e}")
        errors.append(f"Excelファイルの読み込みエラー: {file_path.name}")

    return CheckResult(is_ok=len(errors) == 0, errors=errors)

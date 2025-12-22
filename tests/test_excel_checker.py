"""Tests for excel_checker module."""

from pathlib import Path
import pytest
from openpyxl import Workbook

from src.excel_checker import CheckResult, check_review_record, check_review_checklist


class TestCheckReviewRecord:
    """Test suite for review record checking."""

    def test_check_review_record_all_ok(self, tmp_path: Path) -> None:
        """Test review record with all required fields filled.

        Args:
            tmp_path: Temporary directory path provided by pytest.
        """
        # Create test Excel file
        wb = Workbook()
        ws = wb.active
        ws["AE2"] = "調査レビュー"  # レビュー名称
        ws["AE7"] = "外部太郎"  # 外部メンバ
        ws["AE8"] = "内部花子"  # 内部メンバ

        test_file = tmp_path / "レビュー記録表(社外)_調査_1回目.xlsx"
        wb.save(test_file)

        result = check_review_record(test_file)

        assert result.is_ok is True
        assert len(result.errors) == 0

    def test_check_review_record_missing_name(self, tmp_path: Path) -> None:
        """Test review record with missing review name.

        Args:
            tmp_path: Temporary directory path provided by pytest.
        """
        wb = Workbook()
        ws = wb.active
        # AE2 is missing
        ws["AE7"] = "外部太郎"
        ws["AE8"] = "内部花子"

        test_file = tmp_path / "レビュー記録表(社外)_調査_1回目.xlsx"
        wb.save(test_file)

        result = check_review_record(test_file)

        assert result.is_ok is False
        assert "AE2" in result.errors[0]

    def test_check_review_record_missing_internal_member(self, tmp_path: Path) -> None:
        """Test review record with missing internal member.

        Args:
            tmp_path: Temporary directory path provided by pytest.
        """
        wb = Workbook()
        ws = wb.active
        ws["AE2"] = "調査レビュー"
        ws["AE7"] = "外部太郎"
        # AE8 is missing

        test_file = tmp_path / "レビュー記録表(社外)_調査_1回目.xlsx"
        wb.save(test_file)

        result = check_review_record(test_file)

        assert result.is_ok is False
        assert "AE8" in result.errors[0]

    def test_check_review_record_missing_external_member_for_external(
        self, tmp_path: Path
    ) -> None:
        """Test review record with missing external member (社外 file).

        Args:
            tmp_path: Temporary directory path provided by pytest.
        """
        wb = Workbook()
        ws = wb.active
        ws["AE2"] = "調査レビュー"
        # AE7 is missing
        ws["AE8"] = "内部花子"

        test_file = tmp_path / "レビュー記録表(社外)_調査_1回目.xlsx"
        wb.save(test_file)

        result = check_review_record(test_file)

        assert result.is_ok is False
        assert "AE7" in result.errors[0]

    def test_check_review_record_no_external_member_required_for_internal(
        self, tmp_path: Path
    ) -> None:
        """Test review record without external member requirement (社内 file).

        Args:
            tmp_path: Temporary directory path provided by pytest.
        """
        wb = Workbook()
        ws = wb.active
        ws["AE2"] = "調査レビュー"
        # AE7 is not required for 社内
        ws["AE8"] = "内部花子"

        test_file = tmp_path / "レビュー記録表_調査_社内_1回目.xlsx"
        wb.save(test_file)

        result = check_review_record(test_file)

        assert result.is_ok is True

    def test_check_review_record_permission_error(self, tmp_path: Path) -> None:
        """Test review record with file permission error.

        Args:
            tmp_path: Temporary directory path provided by pytest.
        """
        # Create non-existent file
        test_file = tmp_path / "non_existent.xlsx"

        result = check_review_record(test_file)

        assert result.is_ok is False
        assert len(result.errors) > 0


class TestCheckReviewChecklist:
    """Test suite for review checklist checking."""

    def test_check_review_checklist_all_ok(self, tmp_path: Path) -> None:
        """Test review checklist with all required fields filled.

        Args:
            tmp_path: Temporary directory path provided by pytest.
        """
        wb = Workbook()
        ws = wb.active
        ws["E4"] = "案件名"  # 案件情報
        ws["E5"] = "クォータ"
        ws["E6"] = "アイテム名"
        ws["N6"] = "2025/12/19"  # 日付
        ws["M15"] = "外部太郎"  # 外部メンバ

        test_file = tmp_path / "レビューチェックリスト_030_社外_1回目.xlsx"
        wb.save(test_file)

        result = check_review_checklist(test_file)

        assert result.is_ok is True
        assert len(result.errors) == 0

    def test_check_review_checklist_missing_project_info(self, tmp_path: Path) -> None:
        """Test review checklist with missing project info.

        Args:
            tmp_path: Temporary directory path provided by pytest.
        """
        wb = Workbook()
        ws = wb.active
        # E4 is missing
        ws["E5"] = "クォータ"
        ws["E6"] = "アイテム名"
        ws["N6"] = "2025/12/19"

        test_file = tmp_path / "レビューチェックリスト_030_社内_1回目.xlsx"
        wb.save(test_file)

        result = check_review_checklist(test_file)

        assert result.is_ok is False
        assert "E4" in result.errors[0]

    def test_check_review_checklist_missing_date(self, tmp_path: Path) -> None:
        """Test review checklist with missing date.

        Args:
            tmp_path: Temporary directory path provided by pytest.
        """
        wb = Workbook()
        ws = wb.active
        ws["E4"] = "案件名"
        ws["E5"] = "クォータ"
        ws["E6"] = "アイテム名"
        # N6 is missing

        test_file = tmp_path / "レビューチェックリスト_030_社内_1回目.xlsx"
        wb.save(test_file)

        result = check_review_checklist(test_file)

        assert result.is_ok is False
        assert "N6" in result.errors[0]

    def test_check_review_checklist_missing_external_member_for_external(
        self, tmp_path: Path
    ) -> None:
        """Test review checklist with missing external member (社外 file).

        Args:
            tmp_path: Temporary directory path provided by pytest.
        """
        wb = Workbook()
        ws = wb.active
        ws["E4"] = "案件名"
        ws["E5"] = "クォータ"
        ws["E6"] = "アイテム名"
        ws["N6"] = "2025/12/19"
        # M15 is missing

        test_file = tmp_path / "レビューチェックリスト_030_社外_1回目.xlsx"
        wb.save(test_file)

        result = check_review_checklist(test_file)

        assert result.is_ok is False
        assert "M15" in result.errors[0]

    def test_check_review_checklist_no_external_member_required_for_internal(
        self, tmp_path: Path
    ) -> None:
        """Test review checklist without external member requirement (社内 file).

        Args:
            tmp_path: Temporary directory path provided by pytest.
        """
        wb = Workbook()
        ws = wb.active
        ws["E4"] = "案件名"
        ws["E5"] = "クォータ"
        ws["E6"] = "アイテム名"
        ws["N6"] = "2025/12/19"
        # M15 is not required for 社内

        test_file = tmp_path / "レビューチェックリスト_030_社内_1回目.xlsx"
        wb.save(test_file)

        result = check_review_checklist(test_file)

        assert result.is_ok is True


class TestCheckResult:
    """Test suite for CheckResult dataclass."""

    def test_check_result_ok(self) -> None:
        """Test CheckResult with no errors."""
        result = CheckResult(is_ok=True, errors=[])
        assert result.is_ok is True
        assert len(result.errors) == 0

    def test_check_result_with_errors(self) -> None:
        """Test CheckResult with errors."""
        result = CheckResult(is_ok=False, errors=["Error 1", "Error 2"])
        assert result.is_ok is False
        assert len(result.errors) == 2
